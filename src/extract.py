from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = PROJECT_ROOT / "data" / "Online Retail.xlsx"


def extract(source_file=SOURCE_FILE):
	source_path = Path(source_file)
	if not source_path.exists():
		raise FileNotFoundError(f"Source file not found: {source_path}")

	workbook = load_workbook(source_path, read_only=True, data_only=True)
	worksheet = workbook["Online Retail"]
	rows = worksheet.iter_rows(values_only=True)
	source_columns = [column.lower() for column in next(rows)]
	records = []
	for row_number, row in enumerate(rows, start=1):
		records.append(row)
		if row_number % 100000 == 0:
			print(f"Read {row_number} rows from workbook")
	data = pd.DataFrame(records, columns=source_columns)
	workbook.close()

	required_columns = {
		"invoiceno",
		"stockcode",
		"description",
		"quantity",
		"invoicedate",
		"unitprice",
		"customerid",
		"country",
	}
	if set(data.columns) != required_columns:
		raise ValueError(
			"Unexpected source columns. "
			f"Expected {sorted(required_columns)}, got {sorted(data.columns)}"
		)

	data = data.rename(
		columns={
			"invoiceno": "invoice_no",
			"stockcode": "stock_code",
			"invoicedate": "invoice_date",
			"unitprice": "unit_price",
			"customerid": "customer_id",
		}
	)
	data["customer_id"] = pd.to_numeric(data["customer_id"], errors="coerce").astype("Int64")
	print(f"Rows extracted: {len(data)}")
	return data
